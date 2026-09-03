"""生成 Excel 工具箱的测试文件。

运行后会在项目根目录创建 testdata/ 文件夹，包含覆盖常见场景的 .xlsx 测试文件。
"""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook

# 保证可以 import core
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "testdata")


def _write_workbook(path: str, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r in rows:
            ws.append(r)
    wb.save(path)


def _write_simple(path: str, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in rows:
        ws.append(r)
    wb.save(path)


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"测试文件将生成到: {OUT_DIR}")

    # ========== 1. 合并 - 表头一致,可直接堆叠 ==========
    print("  - 01_sales_q1_same_header.xlsx / 02_sales_q2_same_header.xlsx")
    _write_simple(
        os.path.join(OUT_DIR, "01_sales_q1_same_header.xlsx"),
        [["姓名", "部门", "销售额", "日期"],
         ["张三", "华东", 12000, "2024-01-15"],
         ["李四", "华南", 9800, "2024-02-20"]],
    )
    _write_simple(
        os.path.join(OUT_DIR, "02_sales_q2_same_header.xlsx"),
        [["姓名", "部门", "销售额", "日期"],
         ["王五", "华北", 15000, "2024-04-10"],
         ["赵六", "华东", 13200, "2024-05-05"]],
    )

    # ========== 2. 合并 - 表头不一致,测试差异提示和并集合并 ==========
    print("  - 03_users_with_age.xlsx / 04_users_with_city.xlsx")
    _write_simple(
        os.path.join(OUT_DIR, "03_users_with_age.xlsx"),
        [["姓名", "年龄", "手机号"],
         ["张三", 28, "13800000001"],
         ["李四", 32, "13800000002"]],
    )
    _write_simple(
        os.path.join(OUT_DIR, "04_users_with_city.xlsx"),
        [["姓名", "城市", "手机号"],
         ["王五", "北京", "13800000003"],
         ["赵六", "上海", "13800000004"]],
    )

    # ========== 3. 合并 - 单文件多 Sheet ==========
    print("  - 05_multi_sheets.xlsx")
    _write_workbook(
        os.path.join(OUT_DIR, "05_multi_sheets.xlsx"),
        {
            "1月": [["日期", "收入"], ["2024-01-01", 1000], ["2024-01-02", 1200]],
            "2月": [["日期", "收入"], ["2024-02-01", 900], ["2024-02-02", 1100]],
            "3月": [["日期", "收入"], ["2024-03-01", 1300]],
        },
    )

    # ========== 4. 拆分 - 按行数 ==========
    print("  - 06_big_data_50rows.xlsx")
    rows = [["ID", "姓名", "金额"]]
    for i in range(1, 51):
        rows.append([i, f"客户{i:03d}", i * 100])
    _write_simple(os.path.join(OUT_DIR, "06_big_data_50rows.xlsx"), rows)

    # ========== 5. 拆分 - 按列值,含空值 ==========
    print("  - 07_departments_with_blank.xlsx")
    _write_simple(
        os.path.join(OUT_DIR, "07_departments_with_blank.xlsx"),
        [["部门", "姓名", "金额"],
         ["技术部", "张三", 5000],
         ["技术部", "李四", 6000],
         ["销售部", "王五", 4500],
         [None, "赵六", 3000],        # None
         ["", "钱七", 2800],          # 空字符串
         ["销售部", "孙八", 5500],
         ["运营部", "周九", 4000]],
    )

    # ========== 6. 拆分 - 按 Sheet ==========
    print("  - 08_products_by_category.xlsx")
    _write_workbook(
        os.path.join(OUT_DIR, "08_products_by_category.xlsx"),
        {
            "手机": [["型号", "价格"], ["iPhone15", 5999], ["小米14", 3999]],
            "电脑": [["型号", "价格"], ["MacBook Air", 8999], ["ThinkPad X1", 9999]],
            "平板": [["型号", "价格"], ["iPad", 3299]],
        },
    )

    # ========== 7. 拆分 - 列号超限场景 ==========
    print("  - 09_short_columns.xlsx")
    _write_simple(
        os.path.join(OUT_DIR, "09_short_columns.xlsx"),
        [["姓名", "金额"],
         ["张三", 100],
         ["李四", 200]],
    )

    # ========== 8. 拆分 - 非法文件名场景 ==========
    print("  - 10_invalid_filenames.xlsx")
    _write_simple(
        os.path.join(OUT_DIR, "10_invalid_filenames.xlsx"),
        [["类别", "名称"],
         ["A/B", "包含斜杠"],
         ["C:D", "包含冒号"],
         ["E*F", "包含星号"]],
    )

    print("\n全部生成完毕,共 10 个测试文件。")


if __name__ == "__main__":
    main()
