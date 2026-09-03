"""Excel 合并功能。

支持：
1. 多文件 -> 单文件：多个 Excel 合并成一个工作簿(多 Sheet 或单 Sheet 追加行)
2. 多 Sheet -> 单 Sheet：单个工作簿的所有 Sheet 合并到一个 Sheet
3. 并集表头合并：多文件按表头去重后的并集合并到单 Sheet
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Callable, Iterable, List, Optional

from openpyxl import Workbook, load_workbook


@dataclass
class MergeOptions:
    """合并选项。

    Attributes:
        mode: 合并模式
            - "files_to_sheets": 多个文件 -> 单工作簿,每个文件一个 Sheet
            - "files_to_rows":  多个文件 -> 单工作簿,所有内容追加到一个 Sheet
            - "sheets_to_rows": 单文件 -> 多 Sheet 合并到单 Sheet
            - "files_to_union_rows": 多文件 -> 单工作簿,以所有文件表头的并集作为表头合并
        output_path: 输出文件路径(.xlsx)
        include_header: 追加行模式时,除首个文件外是否包含表头
    """

    mode: str = "files_to_sheets"
    output_path: str = ""
    include_header: bool = True


@dataclass
class HeaderDiff:
    """单个文件相对并集的表头差异。"""

    file_path: str = ""
    file_name: str = ""
    headers: List[str] = field(default_factory=list)
    missing: List[str] = field(default_factory=list)  # 并集有,该文件无


ProgressCallback = Optional[Callable[[int, int, str], None]]


def _emit(cb: ProgressCallback, current: int, total: int, msg: str) -> None:
    if cb:
        try:
            cb(current, total, msg)
        except Exception:
            pass


def _copy_sheet_data(src_ws, dst_ws, include_header: bool, start_row: int) -> int:
    """将 src_ws 的数据复制到 dst_ws,返回写入后的下一可用行号。"""
    rows = src_ws.iter_rows(values_only=True)
    written = start_row
    header_skipped = False
    for r_idx, row in enumerate(rows, start=1):
        if r_idx == 1 and not include_header and not header_skipped:
            header_skipped = True
            continue
        if r_idx == 1 and include_header and start_row > 1:
            # 已经有表头,跳过
            continue
        # 跳过完全空行
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        dst_ws.append(list(row))
        written += 1
    return written


def _read_first_row(file_path: str) -> List[str]:
    """读取指定文件第一个 Sheet 的第一行,空/None 转成空字符串。"""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            return ["" if c is None else str(c) for c in row]
        return []
    finally:
        wb.close()


def scan_header_differences(input_files: Iterable[str]) -> tuple[List[str], List[HeaderDiff]]:
    """扫描多个文件第一个 Sheet 的表头,返回并集表头 + 各文件差异信息。

    Returns:
        (union_headers, diffs)
    """
    files = [f for f in input_files if f and os.path.isfile(f)]
    headers_by_file: dict[str, List[str]] = {}
    for fp in files:
        headers_by_file[fp] = _read_first_row(fp)

    # 求并集,保留首次出现顺序
    seen = set()
    union_headers: List[str] = []
    for fp in files:
        for h in headers_by_file[fp]:
            if h not in seen:
                seen.add(h)
                union_headers.append(h)

    diffs: List[HeaderDiff] = []
    union_set = set(union_headers)
    for fp in files:
        file_headers = headers_by_file[fp]
        file_set = set(file_headers)
        diffs.append(HeaderDiff(
            file_path=fp,
            file_name=os.path.basename(fp),
            headers=file_headers,
            missing=list(union_set - file_set),
        ))

    return union_headers, diffs


def merge_files_to_sheets(
    input_files: Iterable[str],
    output_path: str,
    progress_cb: ProgressCallback = None,
) -> int:
    """多个 Excel 文件 -> 单个工作簿,每个源文件作为一个 Sheet。

    Returns:
        成功合并的 Sheet 数
    """
    files = [f for f in input_files if f and os.path.isfile(f)]
    total = len(files)
    if total == 0:
        raise ValueError("没有可用的输入文件")

    wb_out = Workbook()
    # 删除默认 Sheet
    default_ws = wb_out.active
    wb_out.remove(default_ws)

    for idx, fp in enumerate(files, start=1):
        _emit(progress_cb, idx - 1, total, f"读取: {os.path.basename(fp)}")
        wb_in = load_workbook(fp, data_only=True, read_only=True)
        try:
            for sheet_name in wb_in.sheetnames:
                src = wb_in[sheet_name]
                # 新 Sheet 名称去重
                new_name = sheet_name[:31] if sheet_name else f"Sheet{idx}"
                suffix = 1
                base_name = new_name
                while new_name in wb_out.sheetnames:
                    suffix += 1
                    new_name = f"{base_name[: 31 - len(str(suffix)) - 1]}_{suffix}"
                dst = wb_out.create_sheet(title=new_name)
                for row in src.iter_rows(values_only=True):
                    if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
                        continue
                    dst.append(list(row))
                _emit(
                    progress_cb,
                    idx,
                    total,
                    f"已合并: {os.path.basename(fp)} -> [{new_name}]",
                )
        finally:
            wb_in.close()

    # 至少保证一个 Sheet
    if not wb_out.sheetnames:
        wb_out.create_sheet(title="Sheet1")

    wb_out.save(output_path)
    _emit(progress_cb, total, total, f"完成: {output_path}")
    return len(wb_out.sheetnames)


def merge_files_to_rows(
    input_files: Iterable[str],
    output_path: str,
    include_header: bool = True,
    progress_cb: ProgressCallback = None,
) -> int:
    """多个 Excel 文件 -> 单 Sheet 追加行(首个文件保留表头,其余按需) """
    files = [f for f in input_files if f and os.path.isfile(f)]
    total = len(files)
    if total == 0:
        raise ValueError("没有可用的输入文件")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged"

    next_row = 1
    written_rows = 0
    for idx, fp in enumerate(files, start=1):
        _emit(progress_cb, idx - 1, total, f"读取: {os.path.basename(fp)}")
        wb_in = load_workbook(fp, data_only=True, read_only=True)
        try:
            # 取每个文件第一个 Sheet
            src = wb_in[wb_in.sheetnames[0]]
            before = ws_out.max_row
            # 语义:include_header=True => 仅首个文件保留表头,其余跳过
            #     include_header=False => 所有文件都不写表头
            keep_hdr = include_header and idx == 1
            _copy_sheet_data(src, ws_out, include_header=keep_hdr, start_row=next_row)
            after = ws_out.max_row
            written_rows += max(0, after - before)
            next_row = after + 1 if after >= 1 else 1
        finally:
            wb_in.close()
        _emit(progress_cb, idx, total, f"已追加: {os.path.basename(fp)}")

    wb_out.save(output_path)
    _emit(progress_cb, total, total, f"完成: {output_path} (共 {written_rows} 行)")
    return written_rows


def merge_files_to_union_rows(
    input_files: Iterable[str],
    output_path: str,
    union_headers: List[str],
    progress_cb: ProgressCallback = None,
) -> int:
    """多文件 -> 单 Sheet,以 union_headers 为表头,缺失列补空,数据按行追加。

    每个文件只取第一个 Sheet。
    """
    files = [f for f in input_files if f and os.path.isfile(f)]
    total = len(files)
    if total == 0:
        raise ValueError("没有可用的输入文件")
    if not union_headers:
        raise ValueError("并集表头为空,无法合并")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged"

    # 写入并集表头
    ws_out.append(list(union_headers))

    written_rows = 0
    for idx, fp in enumerate(files, start=1):
        _emit(progress_cb, idx - 1, total, f"读取: {os.path.basename(fp)}")
        wb_in = load_workbook(fp, data_only=True, read_only=True)
        try:
            src = wb_in[wb_in.sheetnames[0]]
            rows_iter = src.iter_rows(values_only=True)
            try:
                file_headers = ["" if c is None else str(c) for c in next(rows_iter)]
            except StopIteration:
                file_headers = []

            # 建立当前文件列名 -> 列索引的映射
            col_index_map = {h: i for i, h in enumerate(file_headers)}

            for row in rows_iter:
                # 跳过完全空行
                if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
                    continue
                new_row: List[Optional[object]] = [None] * len(union_headers)
                for target_idx, header in enumerate(union_headers):
                    src_idx = col_index_map.get(header)
                    if src_idx is not None and src_idx < len(row):
                        new_row[target_idx] = row[src_idx]
                ws_out.append(new_row)
                written_rows += 1
        finally:
            wb_in.close()
        _emit(progress_cb, idx, total, f"已追加: {os.path.basename(fp)}")

    wb_out.save(output_path)
    _emit(progress_cb, total, total, f"完成: {output_path} (共 {written_rows} 行)")
    return written_rows


def merge_sheets_to_rows(
    input_file: str,
    output_path: str,
    include_header: bool = True,
    progress_cb: ProgressCallback = None,
) -> int:
    """单文件多 Sheet -> 单 Sheet 追加行。"""
    if not os.path.isfile(input_file):
        raise ValueError(f"输入文件不存在: {input_file}")

    wb_in = load_workbook(input_file, data_only=True, read_only=True)
    try:
        sheet_names = list(wb_in.sheetnames)
    finally:
        wb_in.close()

    total = len(sheet_names)
    if total == 0:
        raise ValueError("输入文件没有任何 Sheet")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged"

    next_row = 1
    written_rows = 0
    for idx, name in enumerate(sheet_names, start=1):
        _emit(progress_cb, idx - 1, total, f"读取 Sheet: {name}")
        wb_in = load_workbook(input_file, data_only=True, read_only=True)
        try:
            src = wb_in[name]
            before = ws_out.max_row
            keep_hdr = include_header and idx == 1
            _copy_sheet_data(src, ws_out, include_header=keep_hdr, start_row=next_row)
            after = ws_out.max_row
            written_rows += max(0, after - before)
            next_row = after + 1 if after >= 1 else 1
        finally:
            wb_in.close()
        _emit(progress_cb, idx, total, f"已合并 Sheet: {name}")

    wb_out.save(output_path)
    _emit(progress_cb, total, total, f"完成: {output_path} (共 {written_rows} 行)")
    return written_rows


def run_merge(options: MergeOptions, files: List[str], progress_cb: ProgressCallback = None) -> int:
    """统一入口。"""
    if not options.output_path:
        raise ValueError("输出路径不能为空")
    if not options.output_path.lower().endswith(".xlsx"):
        options.output_path += ".xlsx"

    if options.mode == "files_to_sheets":
        return merge_files_to_sheets(files, options.output_path, progress_cb)
    if options.mode == "files_to_rows":
        return merge_files_to_rows(
            files, options.output_path, options.include_header, progress_cb
        )
    if options.mode == "files_to_union_rows":
        union_headers, _ = scan_header_differences(files)
        return merge_files_to_union_rows(
            files, options.output_path, union_headers, progress_cb
        )
    if options.mode == "sheets_to_rows":
        if len(files) != 1:
            raise ValueError("该模式仅支持单个输入文件")
        return merge_sheets_to_rows(
            files[0], options.output_path, options.include_header, progress_cb
        )
    raise ValueError(f"未知合并模式: {options.mode}")
