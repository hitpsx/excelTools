"""Excel 拆分功能。

支持：
1. 按 Sheet 拆分：单文件 -> 多个文件,每个 Sheet 一个文件
2. 按行数拆分：单文件 -> 多个文件,每个文件固定行数
3. 按列值拆分：单文件 -> 多个文件,按指定列的不同值分组
   - 列不存在时抛出友好异常
   - 列值为空/None 的数据统一归到 xxx_blank.xlsx
"""
from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Callable, List, Optional

from openpyxl import Workbook, load_workbook


@dataclass
class SplitOptions:
    """拆分选项。

    Attributes:
        mode: 拆分模式
            - "by_sheet": 按 Sheet 拆分
            - "by_rows": 按行数拆分
            - "by_column": 按列值拆分
        output_dir: 输出目录
        rows_per_file: 按行数拆分时每个文件的行数(含表头)
        column_index: 按列值拆分时的列序号(从 1 开始)
        keep_header: 拆分后是否保留表头
    """

    mode: str = "by_sheet"
    output_dir: str = ""
    rows_per_file: int = 1000
    column_index: int = 1
    keep_header: bool = True


ProgressCallback = Optional[Callable[[int, int, str], None]]


def _emit(cb: ProgressCallback, current: int, total: int, msg: str) -> None:
    if cb:
        try:
            cb(current, total, msg)
        except Exception:
            pass


def _safe_filename(name: str) -> str:
    """过滤 Windows 文件名非法字符。"""
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        name = name.replace(ch, "_")
    name = name.strip().strip(".")
    return name or "unnamed"


def _normalize_group_key(key) -> str:
    """把列值归一化成分组 key;空/None/空字符串 -> blank。"""
    if key is None:
        return "blank"
    if isinstance(key, str):
        s = key.strip()
        return "blank" if s == "" else s
    # 处理 pandas/openpyxl 可能产生的空值
    try:
        if key != key:  # NaN 判断
            return "blank"
    except Exception:
        pass
    return str(key)


def _read_rows(ws) -> List[list]:
    """读取 Sheet 内所有非空行(纯数据,不含合并单元格样式)。"""
    data: List[list] = []
    for row in ws.iter_rows(values_only=True):
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        data.append(list(row))
    return data


def split_by_sheet(
    input_file: str,
    output_dir: str,
    progress_cb: ProgressCallback = None,
) -> int:
    """按 Sheet 拆分 -> 每个 Sheet 一个 xlsx 文件。"""
    if not os.path.isfile(input_file):
        raise ValueError(f"输入文件不存在: {input_file}")
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(input_file))[0]
    wb_in = load_workbook(input_file, data_only=True, read_only=True)
    try:
        sheet_names = list(wb_in.sheetnames)
    finally:
        wb_in.close()

    total = len(sheet_names)
    if total == 0:
        raise ValueError("输入文件没有任何 Sheet")

    count = 0
    for idx, name in enumerate(sheet_names, start=1):
        _emit(progress_cb, idx - 1, total, f"读取 Sheet: {name}")
        wb_in = load_workbook(input_file, data_only=True, read_only=True)
        try:
            src = wb_in[name]
            rows = _read_rows(src)
        finally:
            wb_in.close()

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = name[:31] or "Sheet"
        for r in rows:
            ws_out.append(r)

        out_path = os.path.join(output_dir, f"{base_name}_{_safe_filename(name)}.xlsx")
        wb_out.save(out_path)
        count += 1
        _emit(progress_cb, idx, total, f"已生成: {os.path.basename(out_path)}")

    return count


def split_by_rows(
    input_file: str,
    output_dir: str,
    rows_per_file: int,
    keep_header: bool = True,
    progress_cb: ProgressCallback = None,
) -> int:
    """按行数拆分 -> 每个文件固定行数(含表头时仅首文件含表头)。"""
    if not os.path.isfile(input_file):
        raise ValueError(f"输入文件不存在: {input_file}")
    if rows_per_file <= 0:
        raise ValueError("每个文件的行数必须大于 0")

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_file))[0]

    wb_in = load_workbook(input_file, data_only=True, read_only=True)
    try:
        src = wb_in[wb_in.sheetnames[0]]
        rows = _read_rows(src)
    finally:
        wb_in.close()

    if not rows:
        raise ValueError("输入文件没有可拆分的行")

    header = rows[0] if keep_header else None
    data = rows[1:] if keep_header else rows
    total_batches = (len(data) + rows_per_file - 1) // rows_per_file
    if total_batches == 0:
        total_batches = 1

    count = 0
    for i in range(total_batches):
        start = i * rows_per_file
        end = start + rows_per_file
        chunk = data[start:end]
        if not chunk and i > 0:
            continue

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = f"Part{i + 1}"[:31]
        if header is not None:
            ws_out.append(header)
        for r in chunk:
            ws_out.append(r)

        out_path = os.path.join(output_dir, f"{base_name}_part{i + 1:03d}.xlsx")
        wb_out.save(out_path)
        count += 1
        _emit(progress_cb, i + 1, total_batches, f"已生成: {os.path.basename(out_path)}")

    return count


def split_by_column(
    input_file: str,
    output_dir: str,
    column_index: int,
    keep_header: bool = True,
    progress_cb: ProgressCallback = None,
) -> int:
    """按列值拆分 -> 每个不同的列值生成一个文件。

    列值为空/None/NaN/空字符串的数据统一归到 xxx_blank.xlsx。
    """
    if not os.path.isfile(input_file):
        raise ValueError(f"输入文件不存在: {input_file}")
    if column_index <= 0:
        raise ValueError("列序号必须从 1 开始")

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_file))[0]

    wb_in = load_workbook(input_file, data_only=True, read_only=True)
    try:
        src = wb_in[wb_in.sheetnames[0]]
        rows = _read_rows(src)
    finally:
        wb_in.close()

    if not rows:
        raise ValueError("输入文件没有可拆分的行")
    if column_index > len(rows[0]):
        raise ValueError(
            f"你填的是第 {column_index} 列,但表格里只有 {len(rows[0])} 列,\n"
            f"请检查列号是否填错"
        )

    header = rows[0] if keep_header else None
    data = rows[1:] if keep_header else rows

    groups: dict[str, list] = {}
    for r in data:
        # 防御列数不一致
        raw_key = r[column_index - 1] if len(r) >= column_index else None
        key = _normalize_group_key(raw_key)
        groups.setdefault(key, []).append(r)

    total = len(groups)
    if total == 0:
        total = 1
    count = 0
    for i, (key, items) in enumerate(groups.items(), start=1):
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Data"
        if header is not None:
            ws_out.append(header)
        for r in items:
            ws_out.append(r)

        key_str = _safe_filename(key)
        out_path = os.path.join(output_dir, f"{base_name}_{key_str}.xlsx")
        wb_out.save(out_path)
        count += 1
        _emit(progress_cb, i, total, f"已生成: {os.path.basename(out_path)}")

    return count


def run_split(
    options: SplitOptions,
    input_file: str,
    progress_cb: ProgressCallback = None,
) -> int:
    """统一入口。"""
    if not options.output_dir:
        raise ValueError("输出目录不能为空")

    if options.mode == "by_sheet":
        return split_by_sheet(input_file, options.output_dir, progress_cb)
    if options.mode == "by_rows":
        return split_by_rows(
            input_file,
            options.output_dir,
            options.rows_per_file,
            options.keep_header,
            progress_cb,
        )
    if options.mode == "by_column":
        return split_by_column(
            input_file,
            options.output_dir,
            options.column_index,
            options.keep_header,
            progress_cb,
        )
    raise ValueError(f"未知拆分模式: {options.mode}")
