"""核心合并/拆分逻辑的自测脚本。

生成测试数据 -> 跑合并 -> 跑拆分 -> 断言结果。
"""
import os
import sys
import tempfile

# 保证可以 import core
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook, load_workbook

from core.merger import (
    merge_files_to_sheets,
    merge_files_to_rows,
    merge_sheets_to_rows,
    merge_files_to_union_rows,
    scan_header_differences,
    MergeOptions,
    run_merge,
)
from core.splitter import (
    split_by_sheet,
    split_by_rows,
    split_by_column,
    SplitOptions,
    run_split,
)


def make_sample(path: str, sheet_name: str, rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in rows:
        ws.append(r)
    wb.save(path)


def assert_eq(actual, expected, label: str) -> None:
    if actual != expected:
        raise AssertionError(f"{label} 失败: 期望 {expected!r}, 实际 {actual!r}")
    print(f"  ✓ {label}")


def main() -> None:
    tmp = tempfile.mkdtemp(prefix="excel_tools_test_")
    print(f"[tmp] {tmp}")

    # ---- 准备测试数据 ----
    f1 = os.path.join(tmp, "a.xlsx")
    f2 = os.path.join(tmp, "b.xlsx")
    f3 = os.path.join(tmp, "c.xlsx")
    make_sample(f1, "销售", [["姓名", "金额"], ["张三", 100], ["李四", 200]])
    make_sample(f2, "销售", [["姓名", "金额"], ["王五", 300], ["赵六", 400]])
    make_sample(f3, "其它", [["X", "Y"], [1, 2]])

    # ===== 合并测试 =====
    print("\n[TEST] merge files_to_sheets")
    out1 = os.path.join(tmp, "merge_files_to_sheets.xlsx")
    n = merge_files_to_sheets([f1, f2, f3], out1)
    wb = load_workbook(out1, data_only=True)
    assert_eq(set(wb.sheetnames), {"销售", "销售_2", "其它"}, "Sheet 集合")
    assert_eq(wb["销售"]["A2"].value, "张三", "a.xlsx -> [销售] 数据")
    assert_eq(wb["销售_2"]["A2"].value, "王五", "b.xlsx -> [销售_2] 数据")
    assert_eq(wb["其它"]["A2"].value, 1, "c.xlsx -> [其它] 数据")

    print("\n[TEST] merge files_to_rows")
    out2 = os.path.join(tmp, "merge_files_to_rows.xlsx")
    n = merge_files_to_rows([f1, f2], out2, include_header=False)
    wb = load_workbook(out2, data_only=True)
    ws = wb.active
    # 表头: 张三 100 / 李四 200 / 王五 300 / 赵六 400
    assert_eq(ws.max_row, 4, "files_to_rows 行数")
    assert_eq([ws.cell(1, 1).value, ws.cell(1, 2).value], ["张三", 100], "第 1 行")
    assert_eq([ws.cell(4, 1).value, ws.cell(4, 2).value], ["赵六", 400], "末行")

    print("\n[TEST] merge sheets_to_rows")
    multi = os.path.join(tmp, "multi.xlsx")
    wb = Workbook()
    wb.active.title = "S1"
    wb.active.append(["K", "V"])
    wb.active.append(["k1", "v1"])
    wb.create_sheet("S2").append(["K", "V"])
    wb["S2"].append(["k2", "v2"])
    wb.save(multi)
    out3 = os.path.join(tmp, "merge_sheets_to_rows.xlsx")
    n = merge_sheets_to_rows(multi, out3, include_header=False)
    wb = load_workbook(out3, data_only=True)
    ws = wb.active
    assert_eq(ws.max_row, 2, "sheets_to_rows 行数")
    assert_eq([ws.cell(1, 1).value, ws.cell(1, 2).value], ["k1", "v1"], "首行")
    assert_eq([ws.cell(2, 1).value, ws.cell(2, 2).value], ["k2", "v2"], "末行")

    # ===== 表头差异扫描 + 并集合并 =====
    print("\n[TEST] scan_header_differences")
    diff_a = os.path.join(tmp, "diff_a.xlsx")
    diff_b = os.path.join(tmp, "diff_b.xlsx")
    make_sample(diff_a, "Sheet1", [["姓名", "年龄"], ["张三", 20]])
    make_sample(diff_b, "Sheet1", [["姓名", "城市"], ["李四", "北京"]])
    union_headers, diffs = scan_header_differences([diff_a, diff_b])
    assert_eq(union_headers, ["姓名", "年龄", "城市"], "并集表头")
    by_name = {d.file_name: d for d in diffs}
    assert_eq(set(by_name["diff_b.xlsx"].missing), {"年龄"}, "diff_b 缺失年龄")
    assert_eq(set(by_name["diff_a.xlsx"].missing), {"城市"}, "diff_a 缺失城市")

    print("\n[TEST] merge files_to_union_rows")
    out_union = os.path.join(tmp, "merge_union.xlsx")
    n = merge_files_to_union_rows([diff_a, diff_b], out_union, union_headers)
    wb = load_workbook(out_union, data_only=True)
    ws = wb.active
    assert_eq([ws.cell(1, i).value for i in range(1, 4)], ["姓名", "年龄", "城市"], "并集表头写入")
    assert_eq([ws.cell(2, i).value for i in range(1, 4)], ["张三", 20, None], "diff_a 行缺失城市为 None")
    assert_eq([ws.cell(3, i).value for i in range(1, 4)], ["李四", None, "北京"], "diff_b 行缺失年龄为 None")

    # ===== 拆分测试 =====
    print("\n[TEST] split by_sheet")
    out_dir1 = os.path.join(tmp, "split_sheet")
    n = split_by_sheet(multi, out_dir1)
    assert_eq(n, 2, "by_sheet 文件数")
    files = sorted(os.listdir(out_dir1))
    assert_eq(len(files), 2, "by_sheet 落盘文件数")
    wb = load_workbook(os.path.join(out_dir1, files[0]), data_only=True)
    assert_eq(wb.sheetnames[0], "S1", "by_sheet 文件 1 Sheet 名")

    print("\n[TEST] split by_rows")
    # 造一个 25 行的文件
    big = os.path.join(tmp, "big.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "v"])
    for i in range(1, 26):
        ws.append([i, f"v{i}"])
    wb.save(big)
    out_dir2 = os.path.join(tmp, "split_rows")
    n = split_by_rows(big, out_dir2, rows_per_file=10, keep_header=True)
    assert_eq(n, 3, "by_rows 文件数 (10/10/5)")
    wb1 = load_workbook(os.path.join(out_dir2, "big_part001.xlsx"), data_only=True)
    assert_eq(wb1.active.max_row, 11, "part001 行数(10+1 表头)")
    wb3 = load_workbook(os.path.join(out_dir2, "big_part003.xlsx"), data_only=True)
    assert_eq(wb3.active.max_row, 6, "part003 行数(5+1 表头)")

    print("\n[TEST] split by_column (含空值归到 blank)")
    group = os.path.join(tmp, "group.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["dept", "name"])
    for d, n_ in [("A", "a1"), ("B", "b1"), (None, "n1"), ("", "n2"), ("A", "a2"), ("C", "c1")]:
        ws.append([d, n_])
    wb.save(group)
    out_dir3 = os.path.join(tmp, "split_col")
    n = split_by_column(group, out_dir3, column_index=1, keep_header=True)
    assert_eq(n, 4, "by_column 文件数 (A/B/C/blank)")
    wb_a = load_workbook(os.path.join(out_dir3, "group_A.xlsx"), data_only=True)
    rows_a = [tuple(r) for r in wb_a.active.iter_rows(values_only=True)]
    assert_eq(len(rows_a), 3, "A 组行数 (2+1 表头)")
    wb_blank = load_workbook(os.path.join(out_dir3, "group_blank.xlsx"), data_only=True)
    rows_blank = [tuple(r) for r in wb_blank.active.iter_rows(values_only=True)]
    assert_eq(len(rows_blank), 3, "blank 组行数 (2+1 表头)")

    print("\n[TEST] split by_column 列不存在")
    try:
        split_by_column(group, out_dir3, column_index=5, keep_header=True)
    except ValueError as e:
        assert "你填的是第 5 列" in str(e), f"异常提示不友好: {e}"
        print("  ✓ 列不存在提示友好")

    print("\n[TEST] run_merge / run_split 入口")
    out4 = os.path.join(tmp, "via_opts.xlsx")
    n = run_merge(MergeOptions(mode="files_to_rows", output_path=out4), [f1, f2])
    wb = load_workbook(out4, data_only=True)
    # 默认 include_header=True => 1 表头 + 2(a数据) + 2(b数据) = 5 行
    assert_eq(wb.active.max_row, 5, "run_merge 行数(含表头)")

    out5 = os.path.join(tmp, "via_opts_noheader.xlsx")
    n = run_merge(
        MergeOptions(mode="files_to_rows", output_path=out5, include_header=False),
        [f1, f2],
    )
    wb = load_workbook(out5, data_only=True)
    assert_eq(wb.active.max_row, 4, "run_merge 行数(无表头)")

    out_dir5 = os.path.join(tmp, "split_opts")
    n = run_split(SplitOptions(mode="by_sheet", output_dir=out_dir5), multi)
    assert_eq(n, 2, "run_split 文件数")

    print("\n🎉 全部自测通过")
    print(f"   临时目录: {tmp}")


if __name__ == "__main__":
    main()
